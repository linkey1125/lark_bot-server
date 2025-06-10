import os
import json
import requests
import openpyxl
from flask import Flask, request, Response
from dotenv import load_dotenv
from openai import OpenAI

# 環境変数の読み込み
load_dotenv()
APP_ID = os.environ["LARK_APP_ID"]
APP_SECRET = os.environ["LARK_APP_SECRET"]
OPENAI_API_KEY = os.environ["OPENAI_API_KEY"]

# OpenAIクライアントの初期化
client = OpenAI(api_key=OPENAI_API_KEY)

# Flaskアプリの設定
app = Flask(__name__)

# 処理済みイベントIDを記録するセット（簡易キャッシュ）
processed_event_ids = set()

# Larkのアクセストークンを取得
def get_tenant_access_token():
    url = "https://open.larksuite.com/open-apis/auth/v3/tenant_access_token/internal"
    payload = {"app_id": APP_ID, "app_secret": APP_SECRET}
    headers = {"Content-Type": "application/json"}
    res = requests.post(url, headers=headers, json=payload)
    return res.json().get("tenant_access_token", "")

# Larkにファイルを送信
def send_file_to_lark(chat_id, file_path="output.xlsx"):
    token = get_tenant_access_token()
    upload_url = "https://open.larksuite.com/open-apis/im/v1/files"
    headers = {"Authorization": f"Bearer {token}"}
    files = {"file": open(file_path, "rb")}
    data = {"file_name": os.path.basename(file_path), "file_type": "stream"}
    res = requests.post(upload_url, headers=headers, data=data, files=files)

    res_data = res.json()
    if "data" not in res_data:
        print("ファイルアップロード失敗:", res_data)
        return

    file_key = res_data["data"]["file_key"]
    send_url = "https://open.larksuite.com/open-apis/im/v1/messages?receive_id_type=chat_id"
    payload = {
        "receive_id": chat_id,
        "msg_type": "file",
        "content": json.dumps({"file_key": file_key})
    }
    headers.update({"Content-Type": "application/json"})
    requests.post(send_url, headers=headers, json=payload)

# GPTで案件情報を抽出
def extract_projects_with_gpt(email_text):
    system_prompt = """
あなたはSES営業担当です。以下のメール本文から複数の案件情報を抽出してください。

各案件は以下の形式のJSONにしてください：
{
  "案件名": "",
  "作業内容": "",
  "募集要件": "",
  "募集人数": "",
  "期間": "",
  "勤務場所": "",
  "その他": ""
}

「未記入」や「不明」の項目があっても構いません。複数案件がある場合、JSONのリスト形式で返してください。
"""

    user_prompt = f"以下の案件メール本文を解析してください：\n\n{email_text}"

    try:
        response = client.chat.completions.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.2
        )

        content = response.choices[0].message.content.strip()

        # JSON部分だけ抽出（まれに文章が混ざる場合あり）
        json_start = content.find("[")
        json_end = content.rfind("]") + 1
        json_str = content[json_start:json_end]
        data = json.loads(json_str)

        return data if isinstance(data, list) else [data]

    except Exception as e:
        print("GPT解析失敗:", str(e))
        return []

# Excelファイルへの出力
def export_all_to_excel(data_list, filename="output.xlsx"):
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
    from openpyxl.styles import Alignment, Font, PatternFill

    def clean_value(val):
        if isinstance(val, str) and val.strip() in ["未記入", "不明"]:
            return ""
        return val

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SES案件一覧"

    # ヘッダー
    headers = ['案件名', '作業内容', '募集要件', '募集人数', '期間', '勤務場所', 'その他']
    ws.append(headers)

    # スタイル定義
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # 薄い黄色
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # ヘッダー行スタイル適用
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # データ行を追加
    for data in data_list:
        row = [
            data.get('案件名', ''),
            data.get('作業内容', ''),
            data.get('募集要件', ''),
            data.get('募集人数', ''),
            data.get('期間', ''),
            data.get('勤務場所', ''),
            data.get('その他', '')
        ]
        ws.append(row)

    # オートフィルター
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # 列幅調整（最小幅15、最大幅60）
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(1, len(headers) + 1):
        max_length = max(len(str(ws.cell(row=row, column=col).value or '')) for row in range(1, ws.max_row + 1))
        width = min(max(max_length + 5, 14), 40)
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=width)
    ws.column_dimensions = dim_holder

    # ヘッダー高さ
    ws.row_dimensions[1].height = 20

    # データ行：折り返し表示・行の高さ調整
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 60
        for cell in ws[i]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(filename)



# Webhookエンドポイント
@app.route('/webhook', methods=['POST'])
def lark_webhook():
    data = request.json
    print("Webhook受信データ:", json.dumps(data, indent=4, ensure_ascii=False))

    # challenge応答
    if 'challenge' in data:
        return Response(json.dumps({'challenge': data['challenge']}), status=200, mimetype='application/json')

    # イベントIDを取得し、重複処理を防ぐ
    event_id = data.get('uuid') or data.get('header', {}).get('event_id')
    if event_id in processed_event_ids:
        print(f"重複イベント検出（スキップ）: {event_id}")
        return Response(json.dumps({'status': 'duplicate'}), status=200)
    processed_event_ids.add(event_id)

    try:
        msg_type = data.get('event', {}).get('message', {}).get('message_type')
        if msg_type != 'text':
            print(f"無視されたメッセージタイプ: {msg_type}")
            return Response(json.dumps({'status': 'ignored'}), status=200)

        message = data['event']['message']
        content_str = message.get('content', '{}')
        content_dict = json.loads(content_str)
        email_text = content_dict.get('text', '')

    except Exception as e:
        print(f"メッセージ抽出エラー: {str(e)}")
        return Response(json.dumps({'error': str(e)}), status=400)

    # GPTで案件解析
    parsed_blocks = extract_projects_with_gpt(email_text)

    for i, block in enumerate(parsed_blocks):
        print(f"\n--- Block {i+1} ---")
        print(json.dumps(block, ensure_ascii=False, indent=2))

    # Excel保存 & Lark送信
    export_all_to_excel(parsed_blocks)
    chat_id = message.get("chat_id") or message.get("conversation_id")
    send_file_to_lark(chat_id)

    return Response(json.dumps({'status': 'success'}), status=200)

# アプリ起動
if __name__ == '__main__':
    app.run(debug=True, port=5000)
